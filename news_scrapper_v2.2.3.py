## 2022.03.22. news_scrapper.exe 버전에서 수정
# pip install bs4, openpyxl, lxml -> 패키지 파일의 내용 spec 파일 hiddenimports에 추가!
# v2.0 : 비프음 추가, 대상 언론사 함수 엑셀에 추가.
# v2.1 : 파일 이름 "기사 시작일-종료일_회사 이름" 형태로 수정
# v2.2 : 기사 기간 검색 자동화
# v2.2.1 : main 부분 함수화
# v2.2.3 : 2.2.1 버전에서 수정

### searcher.py
from pickle import TRUE
from urllib import parse
import openpyxl as xl
import datetime
import requests
from bs4 import BeautifulSoup
import os, re
import time     #추가

#비프음 추가(v2.0)
import winsound as sd
def beepsound():
    fr = 2000    # range : 37 ~ 32767
    du = 1000     # 1000 ms ==1second
    sd.Beep(fr, du) # winsound.Beep(frequency, duration)

# 엑셀에서 값 받아오기
def get_query():

    file_path = os.getcwd() #현재 폴더 경로
    dir = f"{file_path}/input.xlsx"
    wb = xl.load_workbook(dir, data_only=TRUE)
    ws = wb['검색조건']     #수정(v2.0) wb.active

    rows = [4,5,6,7,8]
    result = []
    for j in rows:
        result.append(ws.cell(row=j, column=23).value)      #수정(v2.2.3)

    return(result)

##url 설정
def get_search(x, last_date):
    num = x*15+1        #모바일 버전은 페이지 당 15개 기사 표출. 링크 말단에 표출할 기사 순서 기입용.

    data = get_query() 
       
    query = data[0]   #검색어 (\"는 검색창에 큰따옴표 붙여서 넣어주기 위함)
    sort = data[1]                  #관련도순 = 0, 최신순 = 1, 오래된순 = 2

    ##신규 (v2.2.3)-------------------
    if len(last_date) == 10:
        date_start = last_date
    else:
        date_start = data[2]   #시작 날짜: 20xx.xx.xx 형식

    date_end   = data[3]   #끝 날짜: 20xx.xx.xx 형식
    newstype   = data[4]           #유형: 전체 = 0, 포토 = 1, 동영상 = 2, 지면기사 = 3, 보도자료 = 4, 자동생성기사 =5
    
    en_query = parse.quote(query)   #url 인코딩
    url0 = f"https://m.search.naver.com/search.naver?where=m_news&sm=mtb_pge&query={en_query}&sort={sort}&photo={newstype}&field=0&pd=3&ds={date_start}&de={date_end}&mynews=0&office_type=0&office_section_code=0&news_office_checked=&start={num}"

    return url0


###scarpper

def get_newslist(x, last_date): 
            
    res = requests.get(get_search(x,last_date))       ##검색 url 가져오기
    res.raise_for_status()      #문제가 있다면 멈춤
    soup = BeautifulSoup(res.text, "lxml")      #lxml 파서를 통해 re.text를 bs4 객체로 출력

    #ul 태그 중 class 속성이 list_news인 것 찾기
    newslist = soup.find("ul", attrs={"class": "list_news"})
    
    # 검색 결과가 없으면 newslist안에 'a' 태그 "new_tit" class도 없어서 오류 발생. 이때 함수 탈출.
    try:
        titlelist = newslist.find_all('a', attrs={"class":"news_tit"})
    except:
        return
        
    titlelist = newslist.find_all('a', attrs={"class":"news_tit"})  #기사 제목, 링크 긁어오기
    comlist = newslist.find_all('span', attrs={"class":"info"}) #기사 날짜 긁어오기  
    complist = newslist.find_all('a', attrs={"class":"info press"}) #언론사 긁어오기  
    contlist = newslist.find_all('div', attrs={"class":"api_txt_lines dsc_txt"}) # 기사 내용 긁어오기

    # 기사 날짜에서 네이버뉴스 제거해서 리스트 생성
    datenews = []
    for i in range(len(comlist)):
        date1 = comlist[i].get_text()
        if not date1 == "네이버뉴스":
            datenews.append(date1)

    news_scraplist = [] # 리스트 초기화
    for i in range(len(titlelist)):         #한 페이지에서 스크랩한 15개 기사 정보 리스트로 반환
        try:
            title = titlelist[i].get_text()
            link = titlelist[i]["href"]
            news_date = datenews[i]
            press = complist[i].get_text()
            news_cont = contlist[i].get_text()
            numbering = str(f'{x+1}_{i+1}')
            
            news_one = [numbering, title, news_date, press, link, news_cont]
            news_scraplist.append(news_one)    
        except:
            print(f"ERROR{i}")
    print(x*15+1)   #기사 개수 출력

    return(news_scraplist)  

def get_path_name(name):    #파일 경로에 있으면 안되는 문자 제거
    name = re.sub("[\/:*?\"<>|]", "", name)
    return(name)

def get_timeset():
    today = str(datetime.datetime.now())
    timeset = today[5:16]   #02-22 14:12 형식으로 추출
    return(timeset)

#파일명 검수 작업 - 함수화(v2.2)
def get_comp_name():
    # timeset = get_path_name(get_timeset())
    data = get_query()
    str_query = str(data[0])    #엑셀에서 입력한 검색어
    CompName = get_path_name(str_query)   #검색어에서 파일경로에 있으면 안되는 문자 제거
    return(CompName)

##추가 v2.2.3
def get_foldername():
    
    comp_name = get_comp_name()
    fdr_time = get_path_name(get_timeset())
    fdr_name = f"{fdr_time}_{comp_name}"
    return(fdr_name)

### main.py: 엑셀 열고 저장까지(v2.2) ----------------------------------------------------

## main 함수화(v2.2.1)*********************
#시트에 기본 정보 입력
def main(last_date):
    #엑셀 열기
    file_path = os.getcwd() #현재 폴더 경로
    dir = f"{file_path}/input.xlsx" #불러올 input 파일 경로
    wb = xl.load_workbook(dir)

    # 새 워크북 이름 설정 및 ws 지정 (추가 v2.2.2)
    ws = wb.create_sheet('RawData', 1)  #시트 이름 설정
    # wb.remove(wb['검색조건'])           #검색조건 시트 삭제
        # 삭제(v2.2.2)
        #시트이름 설정 및 시트 만들기(수정 v2.2)
        #ws = wb.create_sheet('RawData', 1) #새 시트 만들기 (수정 v2.2.2)

    #메타 데이터 정의(추가 v2.2.2)
    data = get_query() 
    query = data[0]   #검색어 (\"는 검색창에 큰따옴표 붙여서 넣어주기 위함)
    #sort = data[1]         #관련도순 = 0, 최신순 = 1, 오래된순 = 2
    
    ##신규 (v2.2.3)-------------------
    if len(last_date) == 10:
        date_start = last_date
    else:
        date_start = data[2]   #시작 날짜: 20xx.xx.xx 형식

    date_end   = data[3]   #끝 날짜: 20xx.xx.xx 형식
    newstype   = data[4]           #유형: 전체 = 0, 포토 = 1, 동영상 = 2, 지면기사 = 3, 보도자료 = 4, 자동생성기사 =5

    date = datetime.datetime.now()
    url0 = get_search(0, last_date)

    # 추가(v2.2.2)--------------------------
    ## 메타 데이터 엑셀 추가(1~9행)
    ws.append(["", "", "", "", "", "", "", "", "", ""])     #첫 행을 A~J열까지 채워준다
    ws.append(["", "", "", "뉴스 스크래핑 프로그램", "", "", "v2.2.2", "", "", ""])
    ws.append(["", "", "검색어:", query, "", "", "", "", "", ""])          
    ws.append(["", "", "시작 날짜:", date_start, "", "", "", "", "", ""])   
    ws.append(["", "", "끝 날짜:", date_end, "", "", "", "", "", ""])
    ws.append(["", "", "기사 유형:", newstype, "", "", "", "", "", ""])
    ws.append(["", "", "스크랩 일시:", date, "", "", "", "", "", ""])
    ws.append(["", "", "URL:", "", url0, "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "", ""])

    # 수정(v2.1)  "대상언론", "" , 추가
    ws.append(["페이지_번호", "제목", "날짜","언론사", "링크", "내용", "대상언론"]) 

    x =0
    while x < 267:   #267이 최대(4005개)
        try:
            for i in get_newslist(x,last_date):
                ws.append(i) 
        except:
            print(f"Scrap Stop{x}")
            break
        
        x = x +1

    #수정(v2.0)
    # 언론사 함수 추가
    ws['G11'] = "=COUNTIF(대상언론사[대상 언론사],D2)"  #수정(v2.2.2) 'G2'

    # 추가(v2.1)-----------------------
    # 기사 시작일/종료일 가져오기
    max_row = ws.max_row
    print(max_row) #-----------
    first_date = ws.cell(row=11, column=3).value[:10]        #수정(v2.2.2) row=2
    last_date = ws.cell(row=max_row, column=3).value[:10]   #수정(v2.2.1) [:10] 추가

    # 끝 날짜 수정 추가 v2.2.3
    ws['D5'] = last_date

    #동일 파일 확인(v2.2.1)------------
    CompName = get_comp_name()
    fdr_name = get_foldername()
    file_name = f"{first_date}-{last_date}_{CompName}"
    fdr_path = f"{file_path}\\{fdr_name}"
    ful_path = f"{fdr_path}\\{file_name}"
    file_type = ".xlsx"
    
    # 아래 코드 수정 필요!!
    uniq = 1
    while os.path.exists(f"{ful_path}\\{file_type}"):
        wb.save(f"{file_name}({uniq}){file_type}")
        uniq +=1 
    else:
        wb.save(f"{file_name}.xlsx") 
    # --------------------------------

    return(last_date, max_row)   #수정(v2.2.2) 마지막 기사 발간일, 마지막 행 숫자 반환
# -------------------------------------------------------------------------

### 메인 함수 실행*******
# os.mkdir(get_foldername())      #새 폴더 생성

last_date, max_row = main('first')
print(f'메인함수 last_date: {last_date}')
print(f'메인함수 max_row: {max_row}')

# 추가(v2.2.2)

## 최근 기사는 7일 전부터 '7일 전', ... '2일 전'으로 표시. 
# 24시간 이상 48시간 이내는 '1일 전', 24시간 이내는 '23시간 전' 등
# 60분 이내는 '31분 전' 등 표시
# x일 전은 24시간 단위로 보면 되겠다. 발행 날짜보다는 현재 시간으로부터 몇 시간 전인지에 따라 표시.
# last_date는 '20xx.xx.xx', 'x일 전', 'xx분 전', 'x분 전' 네 가지 정도로 표현되겠다.

# 스크랩 기사 개수가 4005개이면 last_date를 검색기간 시작일로 설정해서 다시 스크래핑하자.
# 이때, last_date 길이가 10이 아니면, 이걸 날짜로 변환해주어야 한다. 이건 매우 드문 경우일 것이므로 그냥 확인하라고 할까?

count = 1   #input 파일 수정 횟수
while max_row >= 4015:      #기사가 4005개 모두 채워서 검색되었을 경우
    if len(last_date)==10:   #last_date가 날짜 형태로 반환된 상태라면
        
        last_date, max_row = main(last_date)    #수정된 검색 시작 날짜로부터 다시 기사 스크래핑
        print(f'count:{count}')
        print(f'메인함수 last_date: {last_date}')
        print(f'메인함수 max_row: {max_row}')
        beepsound()
        count += 1
    else:
        break   #last_date가 날짜 형태가 아니면 while 문 종료
time.sleep(1)
#프로그램 종료(삐, 삐, 삐--)
sd.Beep(2000, 700)
time.sleep(0.5)
sd.Beep(2000, 700)
time.sleep(0.5)
sd.Beep(2000, 1500)